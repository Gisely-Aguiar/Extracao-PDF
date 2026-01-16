import pdfplumber
import pandas as pd
import re
import os
from pathlib import Path
import glob
from openpyxl.styles import Font

# ===== CONFIGURAÇÕES =====
PASTA_PDFS = r"C:\Users\Lenovo\Desktop\Gi\freela\extrator_etec\PDFs"  # ALTERE AQUI
SAIDA_EXCEL = "consolidado_etecs.xlsx"

# ===== FUNÇÃO PARA PROCESSAR UM PDF =====
def extrair_dados_pdf(pdf_path):
    """Extrai dados de um único PDF"""
    dados = []
    unidade = ""
    curso = ""
    periodo = ""
    codigo_curso = ""
    
    arquivo_nome = Path(pdf_path).name
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                texto = page.extract_text()
                if not texto:
                    continue
                
                linhas = texto.split("\n")
                
                for linha in linhas:
                    linha = linha.strip()
                    
                    # Capturar UNIDADE
                    if linha.startswith("Unidade:"):
                        unidade_raw = linha.replace("Unidade:", "").strip()
                        # Extrair apenas o nome (remover código E0241.S0000)
                        parts = re.split(r'\s{2,}', unidade_raw)  # Divide por 2+ espaços
                        if len(parts) > 1:
                            unidade = parts[1].strip()
                        else:
                            # Tenta separar código do nome
                            match = re.match(r'^[A-Z]\d{4}\.[A-Z]\d{4}\s+(.+)$', unidade_raw)
                            unidade = match.group(1) if match else unidade_raw
                        
                        # Resetar para nova unidade
                        curso = ""
                        codigo_curso = ""
                        periodo = ""
                    
                    # Capturar CURSO
                    elif "Curso:" in linha:
                        curso_raw = linha.replace("Curso:", "").strip()
                        
                        # Verificar se tem código no início (ex: "2218 ADMINISTRAÇÃO")
                        match_codigo = re.match(r'^(\d+)\s+(.+)$', curso_raw)
                        if match_codigo:
                            codigo_curso = match_codigo.group(1)
                            curso_nome = match_codigo.group(2)
                        else:
                            curso_nome = curso_raw
                        
                        # Verificar se período está na mesma linha
                        if "Período:" in curso_nome:
                            parts = curso_nome.split("Período:", 1)
                            curso = parts[0].strip()
                            periodo = parts[1].strip()
                        else:
                            curso = curso_nome
                    
                    # Capturar PERÍODO em linha separada
                    elif linha.startswith("Período:") and not periodo:
                        periodo = linha.replace("Período:", "").strip()
                    
                    # Pular cabeçalhos das tabelas
                    elif re.match(r'^(CLASS|NOME|INSCRIÇÃO|NOTA|AFRO|ESC|COT|OBS)', linha, re.IGNORECASE):
                        continue
                    elif re.match(r'^\d+\s+\d+', linha):  # Linhas com apenas números
                        continue
                    elif linha.startswith(("LISTA", "ETEC", "Página", "Classif", "Inscrição")):
                        continue
                    
                    # Tentar capturar linha de candidato
                    else:
                        # Padrão 1: com todas as colunas
                        padrao1 = re.match(
                            r'^(\d+)\s+'  # Classificação
                            r'(.+?)\s+'   # Nome
                            r'(E\d{4}\.[A-Z]\d{4}\.\d+-\d+|AUSENTE)\s+'  # Inscrição
                            r'([\d,]+|AUSENTE)\s+'  # Nota
                            r'(SIM|NÃO)\s+'  # Afro
                            r'(SIM|NÃO)\s+'  # Escolaridade
                            r'(SIM|NÃO)\s+'  # COT
                            r'(\d+)\s+'  # C1
                            r'(\d+)\s+'  # C2
                            r'(\d+)\s+'  # C3
                            r'(\d+)\s+'  # C4
                            r'(\d+)\s*'  # C5
                            r'(.*)$',  # Observação
                            linha
                        )
                        
                        # Padrão 2: mais flexível (caso falhe o primeiro)
                        if not padrao1:
                            padrao2 = re.match(
                                r'^(\d+)\s+'  # Classificação
                                r'(.+?)\s+'   # Nome
                                r'(E\d{4}\.[A-Z]\d{4}\.\d+-\d+)\s+'  # Inscrição
                                r'([\d,]+)\s+'  # Nota
                                r'(.+)$',  # Resto
                                linha
                            )
                            
                            if padrao2:
                                # Processar o resto manualmente
                                resto = padrao2.group(5).split()
                                if len(resto) >= 8:  # Afro, Esc, COT, C1-C5, Obs
                                    afro = resto[0] if resto[0] in ['SIM', 'NÃO'] else ''
                                    esc = resto[1] if len(resto) > 1 and resto[1] in ['SIM', 'NÃO'] else ''
                                    cot = resto[2] if len(resto) > 2 and resto[2] in ['SIM', 'NÃO'] else ''
                                    c1 = resto[3] if len(resto) > 3 else ''
                                    c2 = resto[4] if len(resto) > 4 else ''
                                    c3 = resto[5] if len(resto) > 5 else ''
                                    c4 = resto[6] if len(resto) > 6 else ''
                                    c5 = resto[7] if len(resto) > 7 else ''
                                    obs = ' '.join(resto[8:]) if len(resto) > 8 else ''
                                    
                                    padrao1 = type('obj', (object,), {
                                        'group': lambda n: {
                                            1: padrao2.group(1), 2: padrao2.group(2),
                                            3: padrao2.group(3), 4: padrao2.group(4),
                                            5: afro, 6: esc, 7: cot,
                                            8: c1, 9: c2, 10: c3, 11: c4, 12: c5,
                                            13: obs
                                        }.get(n, '')
                                    })()
                        
                        if padrao1:
                            # Formatar curso completo
                            curso_completo = f"{codigo_curso} {curso}".strip() if codigo_curso else curso
                            
                            dados.append([
                                unidade, curso_completo, periodo,
                                padrao1.group(1),  # Classificação
                                padrao1.group(3),  # Inscrição
                                padrao1.group(2),  # Nome
                                padrao1.group(4),  # Nota
                                padrao1.group(5),  # Afro
                                padrao1.group(6),  # Escolaridade
                                padrao1.group(7),  # COT
                                padrao1.group(8),  # C1
                                padrao1.group(9),  # C2
                                padrao1.group(10), # C3
                                padrao1.group(11), # C4
                                padrao1.group(12), # C5
                                padrao1.group(13), # Observação
                                arquivo_nome  # Arquivo de origem
                            ])
        
        return dados
        
    except Exception as e:
        print(f"    ❌ ERRO em {arquivo_nome}: {str(e)}")
        return []

# ===== FUNÇÃO PRINCIPAL =====
def main():
    print("=" * 70)
    print("📚 PROCESSADOR DE PDFs - ETECs (CONSOLIDAÇÃO MULTIPLOS ARQUIVOS)")
    print("=" * 70)
    
    # Verificar pasta
    if not os.path.exists(PASTA_PDFS):
        print(f"\n❌ ERRO: Pasta não encontrada: {PASTA_PDFS}")
        print("   Por favor, ajuste a variável PASTA_PDFS no código.")
        return
    
    # Buscar PDFs
    arquivos_pdf = glob.glob(os.path.join(PASTA_PDFS, "*.pdf"))
    
    if not arquivos_pdf:
        print(f"\n❌ Nenhum arquivo PDF encontrado em: {PASTA_PDFS}")
        print("   Coloque os arquivos PDF na pasta especificada.")
        return
    
    print(f"\n🔍 Encontrados {len(arquivos_pdf)} arquivo(s) PDF:")
    for i, pdf in enumerate(arquivos_pdf, 1):
        print(f"   {i:2}. {Path(pdf).name}")
    
    # Processar cada PDF
    print(f"\n🔄 Processando arquivos...")
    todos_dados = []
    
    for i, pdf_path in enumerate(arquivos_pdf, 1):
        print(f"\n   [{i}/{len(arquivos_pdf)}] {Path(pdf_path).name}")
        dados_pdf = extrair_dados_pdf(pdf_path)
        
        if dados_pdf:
            todos_dados.extend(dados_pdf)
            print(f"      ✅ {len(dados_pdf)} registros extraídos")
        else:
            print(f"      ⚠️  Nenhum registro extraído")
    
    if not todos_dados:
        print(f"\n❌ Nenhum dado foi extraído dos PDFs.")
        return
    
    # Criar DataFrame principal
    colunas = [
        "Unidade", "Curso", "Período", "Classificação", "Inscrição", 
        "Nome", "Nota", "Afro", "Escolaridade", "COT",
        "C1", "C2", "C3", "C4", "C5", "Observação", "Arquivo_Origem"
    ]
    
    df_principal = pd.DataFrame(todos_dados, columns=colunas)
    
    # Converter tipos de dados
    df_principal['Classificação'] = pd.to_numeric(df_principal['Classificação'], errors='coerce')
    for col in ['C1', 'C2', 'C3', 'C4', 'C5']:
        df_principal[col] = pd.to_numeric(df_principal[col], errors='coerce')
    
    # Converter nota para numérico (substituir vírgula por ponto)
    df_principal['Nota_Numerica'] = pd.to_numeric(
        df_principal['Nota'].str.replace(',', '.'), 
        errors='coerce'
    )
    
    # Ordenar
    df_principal = df_principal.sort_values(
        ['Unidade', 'Curso', 'Período', 'Classificação']
    ).reset_index(drop=True)
    
    # ===== CRIAR PLANILHA DE RESUMOS =====
    print(f"\n📊 Criando planilha de resumos...")
    
    # 1. Resumo por Unidade
    resumo_unidade = df_principal.groupby('Unidade').agg({
        'Nome': 'count',
        'Curso': 'nunique',
        'Arquivo_Origem': lambda x: ', '.join(sorted(set(x)))
    }).rename(columns={
        'Nome': 'Total_Candidatos',
        'Curso': 'Cursos_Diferentes',
        'Arquivo_Origem': 'Arquivos_Origem'
    }).reset_index()
    
    # 2. Resumo por Curso/Período
    resumo_curso = df_principal.groupby(['Unidade', 'Curso', 'Período']).agg({
        'Nome': 'count',
        'Nota_Numerica': ['max', 'min', 'mean'],
        'Classificação': 'max'
    }).reset_index()
    
    # Ajustar nomes das colunas multi-level
    resumo_curso.columns = ['Unidade', 'Curso', 'Período', 
                           'Vagas_Preenchidas', 'Maior_Nota', 
                           'Menor_Nota', 'Media_Nota', 'Ultima_Classificacao']
    
    # Arredondar notas
    for col in ['Maior_Nota', 'Menor_Nota', 'Media_Nota']:
        resumo_curso[col] = resumo_curso[col].round(2)
    
    # 3. Resumo Geral
    resumo_geral = pd.DataFrame({
        'Metrica': [
            'Total de Candidatos',
            'Total de Unidades (ETECs)',
            'Total de Cursos Diferentes',
            'Total de Períodos Diferentes',
            'Total de Arquivos PDF Processados',
            'Média de Nota Geral',
            'Maior Nota Encontrada',
            'Menor Nota Encontrada'
        ],
        'Valor': [
            len(df_principal),
            len(df_principal['Unidade'].unique()),
            len(df_principal['Curso'].unique()),
            len(df_principal[['Curso', 'Período']].drop_duplicates()),
            len(df_principal['Arquivo_Origem'].unique()),
            df_principal['Nota_Numerica'].mean().round(2),
            df_principal['Nota_Numerica'].max(),
            df_principal['Nota_Numerica'].min()
        ]
    })
    
    # ===== SALVAR EM EXCEL =====
    print(f"\n💾 Salvando em Excel...")
    
    with pd.ExcelWriter(SAIDA_EXCEL, engine='openpyxl') as writer:
        # Planilha 1: DADOS COMPLETOS
        df_principal.to_excel(writer, index=False, sheet_name='DADOS_COMPLETOS')
        ws1 = writer.sheets['DADOS_COMPLETOS']
        ws1.auto_filter.ref = ws1.dimensions
        
        # Ajustar larguras
        col_widths = {
            'A': 35, 'B': 60, 'C': 25, 'D': 12, 'E': 25,
            'F': 35, 'G': 10, 'H': 8, 'I': 12, 'J': 8,
            'K': 5, 'L': 5, 'M': 5, 'N': 5, 'O': 5,
            'P': 15, 'Q': 30
        }
        
        for col, width in col_widths.items():
            ws1.column_dimensions[col].width = width
        
        ws1.freeze_panes = 'A2'
        
        # Planilha 2: RESUMOS
        ws2 = writer.book.create_sheet('RESUMOS')
        
        # Resumo Geral
        current_row = 1
        ws2.cell(row=current_row, column=1, value="RESUMO GERAL")
        ws2.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 2
        
        # Cabeçalhos do Resumo Geral
        ws2.cell(row=current_row, column=1, value="Métrica")
        ws2.cell(row=current_row, column=2, value="Valor")
        ws2.cell(row=current_row, column=1).font = Font(bold=True)
        ws2.cell(row=current_row, column=2).font = Font(bold=True)
        current_row += 1
        
        # Dados do Resumo Geral
        for idx, row in resumo_geral.iterrows():
            ws2.cell(row=current_row + idx, column=1, value=row['Metrica'])
            ws2.cell(row=current_row + idx, column=2, value=row['Valor'])
        
        current_row += len(resumo_geral) + 3
        
        # Resumo por Unidade
        ws2.cell(row=current_row, column=1, value="RESUMO POR UNIDADE")
        ws2.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 2
        
        # Cabeçalhos do Resumo por Unidade
        for col_idx, col_name in enumerate(resumo_unidade.columns, 1):
            ws2.cell(row=current_row, column=col_idx, value=col_name)
            ws2.cell(row=current_row, column=col_idx).font = Font(bold=True)
        
        # Dados do Resumo por Unidade
        for row_idx, row in resumo_unidade.iterrows():
            for col_idx, value in enumerate(row.values, 1):
                ws2.cell(row=current_row + row_idx + 1, column=col_idx, value=value)
        
        current_row += len(resumo_unidade) + 3
        
        # Resumo por Curso/Período
        ws2.cell(row=current_row, column=1, value="RESUMO POR CURSO/PERÍODO")
        ws2.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 2
        
        # Cabeçalhos do Resumo por Curso/Período
        for col_idx, col_name in enumerate(resumo_curso.columns, 1):
            ws2.cell(row=current_row, column=col_idx, value=col_name)
            ws2.cell(row=current_row, column=col_idx).font = Font(bold=True)
        
        # Dados do Resumo por Curso/Período
        for row_idx, row in resumo_curso.iterrows():
            for col_idx, value in enumerate(row.values, 1):
                ws2.cell(row=current_row + row_idx + 1, column=col_idx, value=value)
        
        # Ajustar larguras na planilha RESUMOS
        for col in range(1, 10):
            col_letter = chr(64 + col) if col <= 26 else chr(64 + (col // 26)) + chr(64 + (col % 26))
            ws2.column_dimensions[col_letter].width = 20
        
        ws2.auto_filter.ref = f"A1:{col_letter}1"
    
    # ===== MOSTRAR ESTATÍSTICAS =====
    print(f"\n✅ PROCESSAMENTO CONCLUÍDO!")
    print("=" * 70)
    print(f"📊 ESTATÍSTICAS FINAIS:")
    print(f"   • Total de registros: {len(df_principal):,}")
    print(f"   • Unidades processadas: {len(df_principal['Unidade'].unique())}")
    print(f"   • Cursos diferentes: {len(df_principal['Curso'].unique())}")
    print(f"   • Arquivos PDF: {len(df_principal['Arquivo_Origem'].unique())}")
    
    print(f"\n🏫 UNIDADES PROCESSADAS:")
    for unidade in sorted(df_principal['Unidade'].unique()):
        qtd = len(df_principal[df_principal['Unidade'] == unidade])
        arquivos = ', '.join(sorted(set(df_principal[df_principal['Unidade'] == unidade]['Arquivo_Origem'])))
        print(f"   • {unidade[:40]:40} → {qtd:3} candidatos")
    
    print(f"\n📚 CURSOS ENCONTRADOS:")
    cursos_unicos = df_principal[['Curso', 'Período']].drop_duplicates()
    for idx, row in cursos_unicos.head(5).iterrows():  # Mostra só 5 primeiros
        print(f"   • {row['Curso'][:40]:40} → {row['Período']}")
    
    if len(cursos_unicos) > 5:
        print(f"   ... e mais {len(cursos_unicos) - 5} cursos")
    
    print(f"\n💾 ARQUIVO SALVO: {SAIDA_EXCEL}")
    print(f"\n📋 O arquivo contém 2 planilhas:")
    print(f"   1. DADOS_COMPLETOS - Todos os candidatos")
    print(f"   2. RESUMOS - Resumo geral, por unidade e por curso")
    
    print(f"\n👁️  AMOSTRA DE DADOS (3 primeiros registros):")
    print("-" * 70)
    amostra = df_principal[['Unidade', 'Curso', 'Período', 'Classificação', 'Nome', 'Nota', 'Arquivo_Origem']].head(3)
    print(amostra.to_string(index=False))

# ===== EXECUÇÃO =====
if __name__ == "__main__":
    # Instruções
    print("\n📝 INSTRUÇÕES DE USO:")
    print("1. Coloque TODOS os PDFs das ETECs em uma pasta")
    print(f"2. No código, altere a variável 'PASTA_PDFS' (linha 14)")
    print("   Exemplo: PASTA_PDFS = r'C:\\Users\\SeuNome\\Downloads\\ETECs'")
    print("3. Execute o script")
    print("\nO script vai:")
    print("   • Processar TODOS os PDFs da pasta")
    print("   • Consolidar tudo em um único Excel")
    print("   • Criar estatísticas e resumos")
    print("-" * 70)
    
    # Perguntar se quer continuar
    resposta = input("\nDeseja executar o script? (S/N): ").strip().upper()
    
    if resposta == 'S':
        main()
    else:
        print("\nExecução cancelada. Ajuste o caminho da pasta e execute novamente.")
    
    print("\n" + "=" * 70)